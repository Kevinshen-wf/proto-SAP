from openpyxl import Workbook
from datetime import datetime

# 创建 Order Report Excel
order_wb = Workbook()
order_ws = order_wb.active
order_ws.title = 'order report'

# Order Report 表头
order_headers = ['Material', 'PO/Line', 'Comments', 'Reply', 'Description']
for col, header in enumerate(order_headers, 1):
    order_ws.cell(row=1, column=col).value = header

# Order Report 数据
order_data = [
    ['MAT001', 'PO001/1', '', '', 'Item 1'],
    ['MAT002', 'PO002/1', '', '', 'Item 2'],
    ['MAT003', 'PO003/2', '', '', 'Item 3'],
]
for row_idx, row_data in enumerate(order_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        order_ws.cell(row=row_idx, column=col_idx).value = value

order_wb.save('pdf_samples/order_report_example.xlsx')
print('✓ Created: pdf_samples/order_report_example.xlsx')

# 创建 WF Closed Excel
source_wb = Workbook()
source_ws = source_wb.active
source_ws.title = 'WF Closed'

# WF Closed 表头
source_headers = ['PO', 'PN ', 'Line', 'PN/Line', 'Description', 'ETA WFSZ ', 'Tracking No', 'Record No']
for col, header in enumerate(source_headers, 1):
    source_ws.cell(row=1, column=col).value = header

# WF Closed 数据
source_data = [
    ['PO001', 'MAT001', '1', 'PO001/1', 'Item 1', datetime(2025, 1, 15), 'TRK-2025-001', 'REC001'],
    ['PO002', 'MAT002', '1', 'PO002/1', 'Item 2', datetime(2025, 1, 20), 'TRK-2025-002', 'REC002'],
    ['PO003', 'MAT003', '2', 'PO003/2', 'Item 3', datetime(2025, 1, 25), 'TRK-2025-003', 'REC003'],
]
for row_idx, row_data in enumerate(source_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        source_ws.cell(row=row_idx, column=col_idx).value = value

source_wb.save('pdf_samples/wf_closed_example.xlsx')
print('✓ Created: pdf_samples/wf_closed_example.xlsx')

print('\n已创建两个测试文件:')
print('1. pdf_samples/order_report_example.xlsx - Order Report 表')
print('2. pdf_samples/wf_closed_example.xlsx - WF Closed 表')
