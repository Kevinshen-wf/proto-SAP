from openpyxl import Workbook
from datetime import datetime

# 创建新的工作簿
wb = Workbook()

# 删除默认的Sheet
wb.remove(wb.active)

# 创建order report工作表
order_report_ws = wb.create_sheet('order report', 0)
order_headers = ['Material', 'PO/Line', 'Comments', 'Reply', 'Description']
for col, header in enumerate(order_headers, 1):
    order_report_ws.cell(row=1, column=col).value = header

# 添加示例数据到order report
order_data = [
    ['MAT001', 'PO001/1', '', '', 'Test Item 1'],
    ['MAT002', 'PO002/1', '', '', 'Test Item 2'],
    ['MAT003', 'PO003/2', '', '', 'Test Item 3'],
]
for row_idx, row_data in enumerate(order_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        order_report_ws.cell(row=row_idx, column=col_idx).value = value

# 创建WF Closed工作表
wf_closed_ws = wb.create_sheet('WF Closed', 1)
wf_headers = ['PO', 'PN ', 'Line', 'PN/Line', 'Description', 'ETA WFSZ ', 'Tracking No', 'Record No']
for col, header in enumerate(wf_headers, 1):
    wf_closed_ws.cell(row=1, column=col).value = header

# 添加示例数据到WF Closed
# 注意：PN/Line应该与order report中的PO/Line匹配
wf_data = [
    ['PO001', 'MAT001', '1', 'PO001/1', 'Test Item 1', datetime(2025, 1, 15), 'TRK001', 'REC001'],
    ['PO002', 'MAT002', '1', 'PO002/1', 'Test Item 2', datetime(2025, 1, 20), 'TRK002', 'REC002'],
    ['PO003', 'MAT003', '2', 'PO003/2', 'Test Item 3', datetime(2025, 1, 25), 'TRK003', 'REC003'],
]
for row_idx, row_data in enumerate(wf_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        wf_closed_ws.cell(row=row_idx, column=col_idx).value = value

# 保存文件
output_path = 'pdf_samples/test_excel_sync.xlsx'
wb.save(output_path)
print(f"测试Excel文件已创建: {output_path}")
