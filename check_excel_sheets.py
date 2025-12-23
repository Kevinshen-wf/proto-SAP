from openpyxl import load_workbook

wb = load_workbook('pdf_samples/order_report_0001000023.xlsx')
print('文件中的工作表列表:')
for i, sheet in enumerate(wb.sheetnames, 1):
    print(f'  {i}. "{sheet}"')
    ws = wb[sheet]
    print(f'     - 数据行数: {ws.max_row - 1}')
    headers = [str(cell.value) for cell in ws[1] if cell.value is not None]
    if headers:
        print(f'     - 列名: {headers[:8]}')
    else:
        print(f'     - 列名: 无')
