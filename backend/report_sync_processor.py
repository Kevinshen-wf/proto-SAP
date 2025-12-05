"""
Report Sync 处理器
用于同步 Excel 报表数据与数据库中的 closed 表数据
"""

import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from psycopg2 import sql


class ReportSyncProcessor:
    """处理 Excel 报表同步"""
    
    def __init__(self, db_manager):
        """
        初始化处理器
        
        Args:
            db_manager: 数据库管理器实例
        """
        self.db_manager = db_manager
    
    def process_excel_sync(self, excel_path):
        """
        处理 Excel 文件同步
        
        Args:
            excel_path: Excel 文件路径
            
        Returns:
            dict: 处理结果
        """
        try:
            # 加载 Excel 工作簿
            wb = load_workbook(excel_path)
            ws = wb.active
            
            # 获取列索引
            headers = [cell.value for cell in ws[1]]
            col_indices = {header: idx + 1 for idx, header in enumerate(headers)}
            
            print(f"[Report Sync] Excel 文件列: {headers}")
            print(f"[Report Sync] 列索引: {col_indices}")
            
            # 验证必要列
            required_columns = ['Material', 'PurchaseOrder', 'Comments', 'Reply', 'Request']
            for col in required_columns:
                if col not in col_indices:
                    return {
                        'success': False,
                        'error': f'缺少必要列: {col}'
                    }
            
            # 处理每一行
            updated_rows = []
            errors = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                try:
                    # 获取关键字段值
                    material = ws.cell(row_idx, col_indices['Material']).value
                    purchase_order = ws.cell(row_idx, col_indices['PurchaseOrder']).value
                    comments_cell = ws.cell(row_idx, col_indices['Comments'])
                    reply_cell = ws.cell(row_idx, col_indices['Reply'])
                    request_cell = ws.cell(row_idx, col_indices['Request'])
                    
                    # 如果 PurchaseOrder 为空，跳过
                    if not purchase_order or not material:
                        continue
                    
                    # 解析 PurchaseOrder 为 PO 和 Line
                    po_info = self._parse_po_line(purchase_order)
                    if not po_info:
                        errors.append(f'行 {row_idx}: 无效的 PurchaseOrder 格式')
                        continue
                    
                    po, line = po_info
                    
                    # 在数据库中检索匹配的 closed 表记录
                    matching_records = self._find_closed_records(po, line, material)
                    
                    print(f"[Report Sync] 行 {row_idx}: 找到 {len(matching_records)} 条匹配记录")
                    if matching_records:
                        for idx, rec in enumerate(matching_records):
                            print(f"[Report Sync]   记录 {idx}: po_line={rec.get('po_line')}, tracking_no={rec.get('tracking_no')}, record_no={rec.get('record_no')}")
                    
                    if not matching_records:
                        errors.append(f'行 {row_idx}: 未找到匹配的 closed 表记录 (PO={po}, Line={line}, PN={material})')
                        continue
                    
                    # 1. 更新 Comments：追加 tracking_no
                    current_comments = comments_cell.value or ''
                    tracking_nos = [r['tracking_no'] for r in matching_records if r.get('tracking_no')]
                    
                    if tracking_nos:
                        new_comments = current_comments
                        for tn in tracking_nos:
                            if tn not in new_comments:
                                new_comments = f"{new_comments}; {tn}".lstrip('; ')
                        comments_cell.value = new_comments
                    
                    # 2. 更新 Reply：从 tracking_no 或 record_no 提取 ETA 日期并加 7 天
                    # 优先从 tracking_no 提取（因为 tracking_no 中可能包含 ETA 信息）
                    eta_source = None
                    eta_date = None
                    
                    # 先尝试从 tracking_no 提取
                    for r in matching_records:
                        tracking_no = r.get('tracking_no')
                        if tracking_no:
                            eta_date = self._extract_and_calculate_eta(tracking_no)
                            if eta_date:
                                eta_source = 'tracking_no'
                                break
                    
                    # 如果 tracking_no 中没找到，再尝试 record_no
                    if not eta_date:
                        record_no_values = [r.get('record_no') for r in matching_records if r.get('record_no')]
                        if record_no_values:
                            eta_date = self._extract_and_calculate_eta(record_no_values[0])
                            if eta_date:
                                eta_source = 'record_no'
                    
                    print(f"[Report Sync] 行 {row_idx}: ETA 来源={eta_source}, 提取的 ETA 日期={eta_date}")
                    
                    if eta_date:
                        reply_cell.value = eta_date
                        print(f"[Report Sync] 行 {row_idx}: 已更新 reply_cell = {eta_date}")
                    else:
                        print(f"[Report Sync] 行 {row_idx}: 无法从 tracking_no 或 record_no 提取 ETA 日期")
                    
                    updated_rows.append({
                        'row': row_idx,
                        'po': po,
                        'line': line,
                        'material': material,
                        'records_count': len(matching_records)
                    })
                    
                except Exception as row_error:
                    errors.append(f'行 {row_idx}: {str(row_error)}')
            
            # 保存修改
            print(f"[Report Sync] 即将保存文件: {excel_path}")
            wb.save(excel_path)
            print(f"[Report Sync] 文件保存成功")
            
            return {
                'success': True,
                'updated_rows': len(updated_rows),
                'details': updated_rows,
                'errors': errors if errors else None,
                'file_path': excel_path
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'处理 Excel 文件失败: {str(e)}'
            }
    
    def _parse_po_line(self, po_line_str):
        """
        解析 PO/Line 字符串
        
        Args:
            po_line_str: 形如 "4500010431/12" 的字符串
            
        Returns:
            tuple: (po, line) 或 None
        """
        if not po_line_str:
            return None
        
        try:
            parts = str(po_line_str).split('/')
            if len(parts) != 2:
                return None
            
            po = parts[0].strip()
            line = parts[1].strip()
            return (po, line)
        except:
            return None
    
    def _find_closed_records(self, po, line, material):
        """
        在 closed 表中查找匹配的记录
        
        Args:
            po: 采购单号
            line: 行号
            material: 物料代码 (PN)
            
        Returns:
            list: 匹配的记录列表
        """
        conn = self.db_manager.get_connection()
        if not conn:
            return []
        
        try:
            cursor = conn.cursor()
            
            # 在 wf_closed 和 non_wf_closed 表中查找
            results = []
            
            for table_name in ['wf_closed', 'non_wf_closed']:
                query = sql.SQL("""
                    SELECT id, po_line, po, pn, tracking_no, record_no, qty
                    FROM purchase_orders.{}
                    WHERE po = %s AND line = %s AND pn = %s
                    ORDER BY id DESC
                """).format(sql.Identifier(table_name))
                
                cursor.execute(query, (po, line, material))
                rows = cursor.fetchall()
                
                if rows:
                    colnames = [desc[0] for desc in cursor.description]
                    for row in rows:
                        record = {colnames[i]: row[i] for i in range(len(colnames))}
                        record['table'] = table_name
                        results.append(record)
            
            cursor.close()
            conn.close()
            return results
            
        except Exception as e:
            print(f"查询 closed 表失败: {e}")
            if conn:
                try:
                    cursor.close()
                    conn.close()
                except:
                    pass
            return []
    
    def _extract_and_calculate_eta(self, record_no_str):
        """
        从 record_no 提取 ETA 日期并加 7 天
        
        Args:
            record_no_str: 形如 "ETA Rotterdam: 1/10/26; Boat Name:CMA CGM PALAIS" 的字符串
            
        Returns:
            str: 计算后的日期字符串 (M/D/YY 格式) 或 None
        """
        if not record_no_str:
            return None
        
        try:
            # 提取 ETA 日期部分：匹配 "ETA xxx: M/D/YY" 的模式
            # 例如：ETA Rotterdam: 1/10/26
            eta_pattern = r'ETA\s+[^:]+:\s*(\d{1,2}/\d{1,2}/\d{2,4})'
            match = re.search(eta_pattern, record_no_str)
            
            if not match:
                return None
            
            date_str = match.group(1)
            
            # 解析日期
            try:
                # 尝试解析 M/D/YY 格式
                date_obj = datetime.strptime(date_str, '%m/%d/%y')
            except ValueError:
                try:
                    # 尝试 M/D/YYYY 格式
                    date_obj = datetime.strptime(date_str, '%m/%d/%Y')
                except ValueError:
                    return None
            
            # 加 7 天
            new_date = date_obj + timedelta(days=7)
            
            # 返回 M/D/YY 格式
            # 处理 %y 可能导致的问题，手动格式化
            month = new_date.month
            day = new_date.day
            year = new_date.year % 100  # 获取最后两位
            return f"{month}/{day}/{year:02d}"
            
        except Exception as e:
            print(f"[Report Sync] 提取 ETA 日期失败: {e}")
            return None


# 全局函数供外部调用
def get_report_sync_processor(db_manager):
    """获取 report sync 处理器"""
    return ReportSyncProcessor(db_manager)
