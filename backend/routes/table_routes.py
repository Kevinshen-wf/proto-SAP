from flask import Blueprint, jsonify, request, send_file
from backend.controllers.table_controller import TableController
from backend.pdf_import_processor import PDFImportProcessor
from backend.report_sync_processor import ReportSyncProcessor
from backend.operation_logger import operation_logger
from backend.models.database import get_db_manager
from backend.utils.jwt_utils import token_required
import os
import tempfile
import io
import json

# 创建蓝图
table_bp = Blueprint('table', __name__, url_prefix='/api')

# 创建控制器实例
table_controller = TableController()
pdf_processor = PDFImportProcessor()

@table_bp.route('/tables/<table_name>', methods=['GET'])
def get_table_data(table_name):
    """获取指定表的数据"""
    result = table_controller.get_table_data(table_name)
    if result['success']:
        return jsonify(result)
    else:
        return jsonify(result), 500

@table_bp.route('/tables/<table_name>/check_duplicates', methods=['POST'])
def check_duplicates(table_name):
    """检查将要插入的数据是否存在主键冲突"""
    try:
        # 获取请求数据
        data_list = request.json.get('data', [])
        result = table_controller.check_duplicates(table_name, data_list)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@table_bp.route('/tables/<table_name>/insert_with_check', methods=['POST'])
def insert_with_check(table_name):
    """插入数据前检查重复并返回结果"""
    try:
        # 获取请求数据
        data_list = request.json.get('data', [])
        
        # 从请求头获取用户邮箱
        user_email = request.headers.get('X-User-Email', 'unknown@example.com')
        
        # 检查重复
        duplicates = table_controller.check_duplicates(table_name, data_list)
        
        # 返回检查结果
        return jsonify({
            'success': True,
            'duplicates': duplicates['duplicates'],
            'data': data_list
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@table_bp.route('/process_pdf', methods=['POST'])
def process_pdf():
    """处理上传的PDF文件"""
    try:
        # 获取上传的文件和公司信息
        file = request.files.get('file')
        company = request.form.get('company')
        
        # 从请求头获取用户邮箱
        user_email = request.headers.get('X-User-Email', 'pdf_importer@example.com')
        
        if not file or not company:
            return jsonify({'success': False, 'error': '缺少文件或公司信息'}), 400
        
        # 创建临时文件
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_file:
            file.save(tmp_file.name)
            tmp_file_path = tmp_file.name
        
        try:
            # 处理PDF文件并检查重复数据
            result = pdf_processor.process_pdf_with_duplicate_check(tmp_file_path, company)
            
            # 注意：这里不再自动插入数据，而是返回处理结果给前端
            # 前端会根据是否有重复数据来决定是否调用插入接口
            # 如果没有重复数据，前端可以直接调用插入接口
            # 如果有重复数据，前端会显示确认对话框，用户确认后再调用插入接口
            
            return jsonify(result)
        finally:
            # 删除临时文件
            if os.path.exists(tmp_file_path):
                os.unlink(tmp_file_path)
                
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@table_bp.route('/insert_data/<table_name>', methods=['POST'])
def insert_data(table_name):
    """插入数据到指定表"""
    try:
        # 获取请求数据
        data_list = request.json.get('data', [])
        
        # 从请求头获取用户邮箱
        user_email = request.headers.get('X-User-Email', 'api_user@example.com')
        
        # 插入数据
        result = pdf_processor.insert_data_with_check(table_name, data_list, user_email)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@table_bp.route('/tables/<table_name>', methods=['PUT'])
def update_row(table_name):
    """更新表中的数据"""
    try:
        # 从查询参数或请求体中获取主键值
        pn = request.args.get('key') or request.json.get('key')
        if not pn:
            return jsonify({'success': False, 'error': '缺少key参数'}), 400
        
        # 获取请求数据，并移除key字段（因为key不是要更新的列）
        updates = request.json.copy() if request.json else {}
        updates.pop('key', None)  # 移除key参数，避免尝试更新key列
        
        # 从请求头获取用户邮箱
        user_email = request.headers.get('X-User-Email', 'unknown@example.com')
        result = table_controller.update_row(table_name, pn, updates, user_email)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@table_bp.route('/tables/<table_name>', methods=['POST'])
def insert_row(table_name):
    """插入新行数据"""
    try:
        # 获取请求数据
        data = request.json
        # 从请求头获取用户邮箱
        user_email = request.headers.get('X-User-Email', 'unknown@example.com')
        result = table_controller.insert_row(table_name, data, user_email)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@table_bp.route('/tables/<table_name>/<key>', methods=['DELETE'])
def delete_row(table_name, key):
    """删除表中的数据"""
    try:
        # 从请求头获取用户邮箱
        user_email = request.headers.get('X-User-Email', 'unknown@example.com')
        # 从查询参数获取关键字段名，如果没有则默认为 'pn'
        key_field = request.args.get('key_field', 'pn')
        result = table_controller.delete_row(table_name, key, user_email, key_field)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@table_bp.route('/tables/<table_name>', methods=['DELETE'])
def delete_row_by_body(table_name):
    """删除表中的数据（使用请求体传递参数）"""
    try:
        # 从请求头获取用户邮箱
        user_email = request.headers.get('X-User-Email', 'unknown@example.com')
        # 从请求体获取 key 和 key_field
        data = request.json or {}
        key = data.get('key')
        key_field = data.get('key_field', 'pn')
        
        if not key:
            return jsonify({'success': False, 'message': '缺少删除键值'}), 400
        
        result = table_controller.delete_row(table_name, key, user_email, key_field)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@table_bp.route('/operation_logs', methods=['GET'])
def get_operation_logs():
    """获取操作日志"""
    try:
        # 从查询参数获取过滤条件
        user_email = request.args.get('user_email')
        table_name = request.args.get('table_name')
        operation = request.args.get('operation')
        limit = request.args.get('limit', 100, type=int)
        
        # 获取操作日志
        logs = operation_logger.get_operation_logs(
            user_email=user_email,
            table_name=table_name,
            operation=operation,
            limit=limit
        )
        
        return jsonify({
            'success': True,
            'data': logs
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@table_bp.route('/report_sync', methods=['POST'])
@token_required
def sync_report():
    """处理 Excel 报表同步，返回修改后的文件"""
    tmp_file_path = None
    try:
        # 获取上传的文件
        file = request.files.get('file')
        
        if not file:
            return jsonify({'success': False, 'error': '缺少上传文件'}), 400
        
        print(f"[Report Sync] 开始处理文件: {file.filename}")
        
        # 创建临时文件
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            file.save(tmp_file.name)
            tmp_file_path = tmp_file.name
        
        print(f"[Report Sync] 临时文件路径: {tmp_file_path}")
        
        try:
            # 获取数据库管理器
            db_manager = get_db_manager()
            
            # 创建处理器并处理 Excel 文件
            processor = ReportSyncProcessor(db_manager)
            print(f"[Report Sync] 开始处理 Excel...")
            result = processor.process_excel_sync(tmp_file_path)
            print(f"[Report Sync] 处理结果: {result['success']}, 更新行数: {result.get('updated_rows', 0)}")
            
            # 如果处理成功，返回修改后的文件
            if result['success']:
                # 读取修改后的文件内容
                print(f"[Report Sync] 读取修改后的文件...")
                with open(tmp_file_path, 'rb') as f:
                    file_data = io.BytesIO(f.read())
                
                # 返回 Excel 文件
                file_data.seek(0)
                response = send_file(
                    file_data,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name='report_synced.xlsx'
                )
                
                # 附带处理结果信息
                # 使用自定义 header 不会影响文件下载
                response.headers['X-Report-Sync-Success'] = 'true'
                response.headers['X-Report-Sync-Updated'] = str(result['updated_rows'])
                if result.get('errors'):
                    response.headers['X-Report-Sync-Errors'] = json.dumps(result['errors'])
                
                print(f"[Report Sync] 成功返回文件")
                return response
            else:
                print(f"[Report Sync] 处理失败: {result.get('error')}")
                return jsonify(result), 400
        except Exception as inner_e:
            print(f"[Report Sync] 内部异常: {str(inner_e)}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'error': str(inner_e)}), 500
    
    except Exception as e:
        print(f"[Report Sync] 外部异常: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    
    finally:
        # 应用下減效应后删除临时文件
        if tmp_file_path and os.path.exists(tmp_file_path):
            try:
                os.unlink(tmp_file_path)
                print(f"[Report Sync] 已删除临时文件: {tmp_file_path}")
            except Exception as cleanup_e:
                print(f"[Report Sync] 删除临时文件失败: {str(cleanup_e)}")

@table_bp.route('/detect_excel_sheets', methods=['POST'])
@token_required
def detect_excel_sheets():
    """检测上传的 Excel 文件中的工作表名称"""
    tmp_file_path = None
    try:
        # 获取上传的文件
        file = request.files.get('file')
        
        if not file:
            return jsonify({'success': False, 'error': '缺少文件'}), 400
        
        print(f"[Excel Detect] 开始检测文件: {file.filename}")
        
        # 创建临时文件
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            file.save(tmp_file.name)
            tmp_file_path = tmp_file.name
        
        try:
            # 检测工作表
            from openpyxl import load_workbook
            wb = load_workbook(tmp_file_path)
            sheet_names = wb.sheetnames
            
            print(f"[Excel Detect] 检测到工作表: {sheet_names}")
            
            return jsonify({
                'success': True,
                'sheets': sheet_names,
                'file_name': file.filename
            })
        
        except Exception as inner_e:
            print(f"[Excel Detect] 检测失败: {str(inner_e)}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'error': str(inner_e)}), 500
    
    except Exception as e:
        print(f"[Excel Detect] 异常: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    
    finally:
        # 删除临时文件
        if tmp_file_path and os.path.exists(tmp_file_path):
            try:
                os.unlink(tmp_file_path)
                print(f"[Excel Detect] 已删除临时文件: {tmp_file_path}")
            except Exception as cleanup_e:
                print(f"[Excel Detect] 删除临时文件失败: {str(cleanup_e)}")

@table_bp.route('/excel_sync', methods=['POST'])
@token_required
def sync_excel():
    """从上传的两个 Excel 文件同步数据到 order report"""
    tmp_order_file_path = None
    tmp_source_file_path = None
    try:
        # 获取上传的文件
        order_report_file = request.files.get('order_report_file')
        source_file = request.files.get('source_file')
        source_sheet = request.form.get('source_sheet', 'WF Closed')
        order_sheet = request.form.get('order_sheet', 'Order Report')
        
        if not order_report_file:
            return jsonify({'success': False, 'error': '缺少 Order Report 文件'}), 400
        
        if not source_file:
            return jsonify({'success': False, 'error': '缺少源数据文件'}), 400
        
        print(f"[Excel Sync] 开始处理文件: {order_report_file.filename} 和 {source_file.filename}")
        print(f"[Excel Sync] 源表: {source_sheet}, Order sheet: {order_sheet}")
        
        # 创建临时文件
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            order_report_file.save(tmp_file.name)
            tmp_order_file_path = tmp_file.name
        
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            source_file.save(tmp_file.name)
            tmp_source_file_path = tmp_file.name
        
        print(f"[Excel Sync] 临时文件路径: {tmp_order_file_path} 和 {tmp_source_file_path}")
        
        try:
            # 创建处理器并处理 Excel 文件
            processor = ExcelSyncProcessor()
            print(f"[Excel Sync] 开始处理 Excel...")
            result = processor.process_excel_sync_two_files(
                tmp_order_file_path,
                tmp_source_file_path,
                order_sheet=order_sheet,
                source_sheet=source_sheet
            )
            print(f"[Excel Sync] 处理结果: {result['success']}, 更新行数: {result.get('updated_rows', 0)}")
            
            # 如果处理成功，返回修改后的 order report 文件
            if result['success']:
                # 读取修改后的文件内容
                print(f"[Excel Sync] 读取修改后的 order report 文件...")
                with open(tmp_order_file_path, 'rb') as f:
                    file_data = io.BytesIO(f.read())
                
                # 返回 Excel 文件
                file_data.seek(0)
                response = send_file(
                    file_data,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name='order_report_synced.xlsx'
                )
                
                # 附带处理结果信息
                response.headers['X-Excel-Sync-Success'] = 'true'
                response.headers['X-Excel-Sync-Updated'] = str(result['updated_rows'])
                if result.get('errors'):
                    response.headers['X-Excel-Sync-Errors'] = json.dumps(result['errors'])
                
                print(f"[Excel Sync] 成功返回文件")
                return response
            else:
                print(f"[Excel Sync] 处理失败: {result.get('error')}")
                return jsonify(result), 400
        except Exception as inner_e:
            print(f"[Excel Sync] 内部异常: {str(inner_e)}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'error': str(inner_e)}), 500
    
    except Exception as e:
        print(f"[Excel Sync] 外部异常: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    
    finally:
        # 删除临时文件
        if tmp_order_file_path and os.path.exists(tmp_order_file_path):
            try:
                os.unlink(tmp_order_file_path)
                print(f"[Excel Sync] 已删除临时文件: {tmp_order_file_path}")
            except Exception as cleanup_e:
                print(f"[Excel Sync] 删除临时文件失败: {str(cleanup_e)}")
        
        if tmp_source_file_path and os.path.exists(tmp_source_file_path):
            try:
                os.unlink(tmp_source_file_path)
                print(f"[Excel Sync] 已删除临时文件: {tmp_source_file_path}")
            except Exception as cleanup_e:
                print(f"[Excel Sync] 删除临时文件失败: {str(cleanup_e)}")