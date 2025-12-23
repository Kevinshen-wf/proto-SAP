"""
Excel 同步处理器
从上传的 Excel 中读取 WF Closed 表数据，
根据 order report 中的数据进行匹配和更新
"""

import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class ExcelSyncProcessor:
    """处理从 Excel 源表同步数据到 order report"""
    
    def __init__(self):
        pass
    
    def process_excel_sync_two_files(self, order_report_path, source_data_path, order_sheet='Order Report', source_sheet='WF Closed'):
        """
        从两个独立的Excel文件同步数据
        
        Args:
            order_report_path: order report Excel 文件路径
            source_data_path: 源数据 Excel 文件路径
            order_sheet: order report 工作表名称
            source_sheet: 源数据工作表名称（WF Closed 或 Non-WF Closed）
            
        Returns:
            dict: 处理结果
        """
        try:
            # 加载两个工作簿
            order_wb = load_workbook(order_report_path)
            source_wb = load_workbook(source_data_path)
            
            # 检查必要的工作表
            if order_sheet not in order_wb.sheetnames:
                return {
                    'success': False,
                    'error': f'缺少工作表 (Order Report 文件): {order_sheet}'
                }
            
            if source_sheet not in source_wb.sheetnames:
                return {
                    'success': False,
                    'error': f'缺少工作表 (Source 文件): {source_sheet}'
                }
            
            # 加载工作表
            order_ws = order_wb[order_sheet]
            source_ws = source_wb[source_sheet]
            
            # 解析源表数据
            source_data = self._parse_source_sheet(source_ws)
            if not source_data:
                return {
                    'success': False,
                    'error': '无法从源表读取数据'
                }
            
            print(f"[Excel Sync] 从 {source_sheet} 读取了 {len(source_data)} 条记录")
            
            # 处理 order report
            updated_rows = []
            errors = []
            
            # 获取 order report 的列索引
            order_headers = [cell.value for cell in order_ws[1]]
            col_indices = {header: idx + 1 for idx, header in enumerate(order_headers)}
            
            print(f"[Excel Sync] Order report 列: {order_headers}")
            print(f"[Excel Sync] 列索引: {col_indices}")
            
            # 检查必要列
            required_cols = ['Material', 'PurchaseOrder', 'Reply']
            missing_cols = [col for col in required_cols if col not in col_indices]
            
            if missing_cols:
                return {
                    'success': False,
                    'error': f'Order report 缺少列: {", ".join(missing_cols)}'
                }
            
            # 检查或创建 Comments 列
            if 'Comments' not in col_indices:
                # 在 Reply 列之后添加 Comments 列
                comments_col_idx = len(order_headers) + 1
                order_ws.cell(1, comments_col_idx).value = 'Comments'
                col_indices['Comments'] = comments_col_idx
                print(f"[Excel Sync] 创建了 Comments 列在列 {comments_col_idx}")
            else:
                comments_col_idx = col_indices['Comments']
                print(f"[Excel Sync] 使用已存在的 Comments 列在列 {comments_col_idx}")
            
            # 处理 order report 中的每一行
            for row_idx in range(2, order_ws.max_row + 1):
                try:
                    material = order_ws.cell(row_idx, col_indices['Material']).value
                    purchase_order = order_ws.cell(row_idx, col_indices['PurchaseOrder']).value
                    reply_cell = order_ws.cell(row_idx, col_indices['Reply'])
                    comments_cell = order_ws.cell(row_idx, col_indices['Comments'])
                    
                    if not purchase_order or not material:
                        continue
                    
                    # 解析 PurchaseOrder 为 PO 和 Line
                    po_info = self._parse_po_line(purchase_order)
                    if not po_info:
                        errors.append(f'行 {row_idx}: 无效的 PurchaseOrder 格式: {purchase_order}')
                        continue
                    
                    po, line = po_info
                    
                    # 在源表中查找匹配的记录
                    matching_record = self._find_matching_record_by_po_line(source_data, po, line, material)
                    
                    if matching_record:
                        # 1. 更新 Comments：追加 Tracking No（如果存在）
                        tracking_no = matching_record.get('Tracking No')
                        if tracking_no:
                            current_comments = comments_cell.value or ''
                            tracking_no_str = str(tracking_no).strip()
                            if tracking_no_str not in str(current_comments):
                                new_comments = f"{current_comments}; {tracking_no_str}".lstrip('; ')
                                comments_cell.value = new_comments
                                print(f"[Excel Sync] 行 {row_idx}: 已更新 Comments 为 {new_comments}")
                        
                        # 2. 更新 Reply：从 ETA 日期 +7 天
                        eta_date = self._get_eta_date(matching_record)
                        if eta_date:
                            reply_str = eta_date.strftime('%m/%d/%y')
                            reply_cell.value = reply_str
                            print(f"[Excel Sync] 行 {row_idx}: 已更新 Reply 为 {reply_str}")
                        
                        updated_rows.append({
                            'row': row_idx,
                            'po': po,
                            'line': line,
                            'material': material
                        })
                    else:
                        print(f"[Excel Sync] 行 {row_idx}: 未找到匹配的源数据 (PO={po}, Line={line}, Material={material})")
                
                except Exception as row_error:
                    errors.append(f'行 {row_idx}: {str(row_error)}')
            
            # 保存修改后的 order report 文件
            print(f"[Excel Sync] 即将保存 order report 文件: {order_report_path}")
            order_wb.save(order_report_path)
            print(f"[Excel Sync] 文件保存成功")
            
            return {
                'success': True,
                'updated_rows': len(updated_rows),
                'details': updated_rows,
                'errors': errors if errors else None,
                'file_path': order_report_path
            }
        
        except Exception as e:
            return {
                'success': False,
                'error': f'处理 Excel 文件失败: {str(e)}'
            }
    
    def _parse_source_sheet(self, worksheet):
        """解析源表数据"""
        try:
            headers = [cell.value for cell in worksheet[1]]
            data = []
            
            for row_idx in range(2, worksheet.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers, 1):
                    cell_value = worksheet.cell(row_idx, col_idx).value
                    # 该序列中的 key 不仅是 header 本身，也包括接死的版本（有/没有尾部空格）
                    row_data[header] = cell_value
                    # 如果 header 有尾部空格，也可以用不带空格的版本查找
                    if header and isinstance(header, str) and header != header.rstrip():
                        row_data[header.rstrip()] = cell_value
                
                # 只保留有数据的行
                if any(v is not None for v in row_data.values()):
                    data.append(row_data)
            
            return data
        except Exception as e:
            print(f"解析源表失败: {e}")
            return []
    
    def _find_matching_record(self, source_data, po_line, material):
        """在源数据中查找匹配的记录"""
        for record in source_data:
            # 比较 PO/Line（考虑空格和大小写）
            source_po_line = record.get('PN/Line') or record.get('PO/Line')
            if source_po_line:
                source_po_line = str(source_po_line).strip()
            
            target_po_line = str(po_line).strip() if po_line else None
            
            # 比较 PN（物料号）
            source_pn = record.get('PN') or record.get('PN ')
            target_pn = str(material).strip() if material else None
            
            # 匹配逻辑：可以通过 PO + Line 或物料号匹配
            if source_po_line == target_po_line:
                return record
            
            # 尝试从 PO/Line 中提取信息
            if target_po_line and '/' in str(target_po_line):
                parts = str(target_po_line).split('/')
                if len(parts) == 2:
                    target_po = parts[0].strip()
                    # 检查源数据的 PO
                    source_po = record.get('PO')
                    if source_po and str(source_po).strip() == target_po:
                        return record
        
        return None
    
    def _parse_po_line(self, po_line_str):
        """
        解析 PO/Line 字符串
        
        Args:
            po_line_str: 形如 "4500010431/12" 的字符串
            
        Returns:
            tuple: (po, line) 或 None
        """
        if not po_line_str:
            return None
        
        try:
            parts = str(po_line_str).split('/')
            if len(parts) != 2:
                return None
            
            po = parts[0].strip()
            line = parts[1].strip()
            return (po, line)
        except:
            return None
    
    def _find_matching_record_by_po_line(self, source_data, po, line, material):
        """
        根据 PO、Line 和 Material 在源数据中查找匹配的记录
        
        Args:
            source_data: 源表数据列表
            po: 采购单号
            line: 行号
            material: 物料代码
            
        Returns:
            dict: 匹配的记录或 None
        """
        for record in source_data:
            # 尝试从 PN/Line 或 PO/Line 列提取信息
            source_po_line = record.get('PN/Line') or record.get('PO/Line')
            if source_po_line:
                source_po_line = str(source_po_line).strip()
                # 解析 PO 和 Line
                parts = source_po_line.split('/')
                if len(parts) == 2:
                    source_po = parts[0].strip()
                    source_line = parts[1].strip()
                    source_pn = record.get('PN') or record.get('PN ')
                    
                    # 匹配 PO、Line 和 Material
                    if source_po == po and source_line == line and str(source_pn).strip() == str(material).strip():
                        return record
        
        return None
    
    def _get_eta_date(self, record):
        """从记录中获取 ETA 日期并加 7 天
        
        Returns:
            datetime: ETA + 7 天的日期对象或 None
        """
        # 优先使用 ETA WFSZ 字段
        eta_wfsz = record.get('ETA WFSZ')
        print(f"[Excel Sync] _get_eta_date: 查找 ETA WFSZ, 值={eta_wfsz}")
        if eta_wfsz:
            if isinstance(eta_wfsz, datetime):
                result = eta_wfsz + timedelta(days=7)
                print(f"[Excel Sync] _get_eta_date: ETA WFSZ 是 datetime 对象, 返回 {result}")
                return result
            try:
                eta_date = datetime.strptime(str(eta_wfsz), '%Y-%m-%d')
                result = eta_date + timedelta(days=7)
                print(f"[Excel Sync] _get_eta_date: ETA WFSZ 解析成功, 返回 {result}")
                return result
            except Exception as e:
                print(f"[Excel Sync] _get_eta_date: ETA WFSZ 解析失败: {e}")
        
        # 尝试从 Record No 中提取 ETA
        record_no = record.get('Record No')
        print(f"[Excel Sync] _get_eta_date: 查找 Record No, 值={record_no}")
        if record_no:
            eta_pattern = r'ETA\s+[^:]+:\s*(\d{1,2}/\d{1,2}/\d{2,4})'
            match = re.search(eta_pattern, str(record_no))
            if match:
                date_str = match.group(1)
                print(f"[Excel Sync] _get_eta_date: 从 Record No 提取到 ETA 日期 {date_str}")
                try:
                    eta_date = datetime.strptime(date_str, '%m/%d/%y')
                    result = eta_date + timedelta(days=7)
                    print(f"[Excel Sync] _get_eta_date: 日期解析成功 (%m/%d/%y), 返回 {result}")
                    return result
                except ValueError:
                    try:
                        eta_date = datetime.strptime(date_str, '%m/%d/%Y')
                        result = eta_date + timedelta(days=7)
                        print(f"[Excel Sync] _get_eta_date: 日期解析成功 (%m/%d/%Y), 返回 {result}")
                        return result
                    except Exception as e:
                        print(f"[Excel Sync] _get_eta_date: 日期解析失败: {e}")
        
        print(f"[Excel Sync] _get_eta_date: 无法从该记录提取 ETA 日期")
        return None


def get_excel_sync_processor():
    """获取 Excel 同步处理器"""
    return ExcelSyncProcessor()
