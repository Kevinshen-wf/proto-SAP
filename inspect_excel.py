#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook

wb = load_workbook('pdf_samples/Intercompany Shipment Tracking.xlsx')
print("Sheet names:", wb.sheetnames)

# 检查WF Closed表
if 'WF Closed' in wb.sheetnames:
    ws = wb['WF Closed']
    print("\nWF Closed sheet - 列:")
    headers = [cell.value for cell in ws[1]]
    print(headers)
    print("\n前5行数据:")
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True)):
        print(f"行{i+2}: {row}")
